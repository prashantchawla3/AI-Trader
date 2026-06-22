#!/usr/bin/env python3
"""
Stage 2 — Local transcript extraction for the DaytradeWarrior video list.

Reads URLs from DaytradeWarrior_Video_List.xlsx, downloads English subtitles
(manual or auto) with yt-dlp, cleans them to plain text, and writes:
  - transcripts/<video_id>.txt        one file per video
  - transcripts_combined.jsonl        one JSON object per line (id, title, url, text)
  - transcript_progress.csv           status log so the run is fully resumable

Resumable: rerun anytime; already-completed videos are skipped.

Requirements (install once):
    pip install yt-dlp openpyxl
    # A JS runtime improves reliability. Install deno: https://docs.deno.com/runtime/
    # On most systems:  curl -fsSL https://deno.land/install.sh | sh

Usage:
    python extract_transcripts.py
    python extract_transcripts.py --limit 50        # test on first 50
    python extract_transcripts.py --sleep 2.0       # be gentler on YouTube
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import time

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "DaytradeWarrior_Video_List.xlsx")
OUT_DIR = os.path.join(HERE, "transcripts")
COMBINED = os.path.join(HERE, "transcripts_combined.jsonl")
PROGRESS = os.path.join(HERE, "transcript_progress.csv")


def load_videos():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, read_only=True)
    ws = wb["Video List"]
    rows = []
    header = None
    for r in ws.iter_rows(values_only=True):
        if header is None:
            header = r
            continue
        idx = {h: i for i, h in enumerate(header)}
        rows.append({
            "title": r[idx["Title"]],
            "url": r[idx["URL"]],
            "id": r[idx["Video ID"]],
        })
    return rows


def vtt_to_text(path):
    text = open(path, encoding="utf-8", errors="ignore").read()
    out, seen = [], set()
    for ln in text.splitlines():
        if ("-->" in ln or ln.strip().isdigit() or ln.startswith("WEBVTT")
                or ln.startswith(("Kind:", "Language:")) or not ln.strip()):
            continue
        ln = re.sub(r"<[^>]+>", "", ln).strip()
        ln = ln.replace("&amp;", "&").replace("&#39;", "'").replace("&quot;", '"')
        if ln and ln not in seen:
            seen.add(ln)
            out.append(ln)
    return " ".join(out)


def fetch_one(video_id, url, sleep):
    """Return cleaned transcript text, or None if unavailable."""
    with tempfile.TemporaryDirectory() as tmp:
        out_tmpl = os.path.join(tmp, "%(id)s")
        cmd = [
            "yt-dlp", "--skip-download",
            "--write-auto-sub", "--write-sub",
            "--sub-lang", "en", "--sub-format", "vtt",
            "--retries", "3", "--no-warnings", "--quiet",
            "-o", out_tmpl, url,
        ]
        # If your network does TLS interception (corporate proxy / sandbox),
        # uncomment the next line:
        # cmd.insert(1, "--no-check-certificates")
        try:
            subprocess.run(cmd, check=False, timeout=120,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except subprocess.TimeoutExpired:
            return None
        vtts = [f for f in os.listdir(tmp) if f.endswith(".vtt")]
        if not vtts:
            return None
        # Prefer a manual en.vtt over auto if both exist
        vtts.sort(key=lambda f: (".en.vtt" not in f))
        time.sleep(sleep)
        return vtt_to_text(os.path.join(tmp, vtts[0]))


def load_done():
    done = set()
    if os.path.exists(PROGRESS):
        with open(PROGRESS) as f:
            for row in csv.reader(f):
                if row and row[1] == "ok":
                    done.add(row[0])
    return done


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="process only first N (0 = all)")
    ap.add_argument("--sleep", type=float, default=1.0, help="seconds to wait between videos")
    args = ap.parse_args()

    os.makedirs(OUT_DIR, exist_ok=True)
    videos = load_videos()
    if args.limit:
        videos = videos[:args.limit]
    done = load_done()

    prog = open(PROGRESS, "a", newline="")
    pw = csv.writer(prog)
    comb = open(COMBINED, "a", encoding="utf-8")

    total = len(videos)
    ok = skip = fail = 0
    for i, v in enumerate(videos, 1):
        vid = v["id"]
        if vid in done:
            skip += 1
            continue
        text = fetch_one(vid, v["url"], args.sleep)
        if text:
            with open(os.path.join(OUT_DIR, f"{vid}.txt"), "w", encoding="utf-8") as f:
                f.write(text)
            comb.write(json.dumps({"id": vid, "title": v["title"],
                                   "url": v["url"], "text": text}) + "\n")
            comb.flush()
            pw.writerow([vid, "ok", len(text)])
            ok += 1
            status = f"ok ({len(text)} chars)"
        else:
            pw.writerow([vid, "no_transcript", 0])
            fail += 1
            status = "NO TRANSCRIPT"
        prog.flush()
        print(f"[{i}/{total}] {vid}  {status}  | ok={ok} fail={fail} skip={skip}",
              flush=True)

    prog.close()
    comb.close()
    print(f"\nDone. ok={ok}  no_transcript={fail}  skipped={skip}")
    print(f"Per-video text:  {OUT_DIR}/")
    print(f"Combined corpus: {COMBINED}")


if __name__ == "__main__":
    main()
