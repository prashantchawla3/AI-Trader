#!/usr/bin/env python3
"""
Transcript extractor for the AI/ML trading video list.

Reads video IDs from ai_trading_youtube_tutorials.xlsx (column C), then for each
video tries, in order:
  1. youtube-transcript-api  -> existing captions (manual or auto), fastest
  2. yt-dlp auto-subtitles   -> fallback caption source
  3. Whisper (faster-whisper)-> transcribe audio when NO captions exist
Outputs ONE plain-text file per video (no timestamps) + a manifest.csv.
Fully RESUMABLE: re-running skips anything already completed.

------------------------------------------------------------------------------
WHY RUN THIS LOCALLY (not in a cloud sandbox):
YouTube blocks datacenter IPs for bulk caption access. On your normal home/work
connection this works fine. If you later hit blocks at scale, add a proxy (see
PROXY below) or slow down DELAY.
------------------------------------------------------------------------------

SETUP:
  pip install youtube-transcript-api yt-dlp faster-whisper openpyxl
  # Whisper fallback also needs ffmpeg installed on your system:
  #   macOS:  brew install ffmpeg
  #   Ubuntu: sudo apt install ffmpeg
  #   Windows: https://www.gyan.dev/ffmpeg/builds/  (add to PATH)

RUN:
  python extract_transcripts.py
  # process only the first N (handy for a test run):
  python extract_transcripts.py --limit 10
  # skip the Whisper step entirely (captions only):
  python extract_transcripts.py --no-whisper
"""

import argparse, csv, os, random, re, sys, time, glob

# ----------------------------- CONFIG ---------------------------------------
XLSX_PATH   = "ai_trading_youtube_tutorials.xlsx"   # input spreadsheet
OUT_DIR     = "transcripts"                         # one .txt per video
MANIFEST    = "manifest.csv"                         # status log
DELAY       = (1.0, 3.0)     # random sleep (sec) between videos to avoid blocks
RETRIES     = 3              # retries per caption attempt
WHISPER_MODEL = "small"     # tiny|base|small|medium|large-v3 (bigger = better/slower)
PROXY       = None          # e.g. "http://user:pass@host:port" if you get blocked
# ----------------------------------------------------------------------------

def log(*a): print(*a, flush=True)

def read_ids(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb["YouTube Tutorials"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # columns: #, Title, URL, Channel, Category, Skill, Length, Free/Paid, Notes
        url = r[2] if len(r) > 2 else None
        if not url or "watch?v=" not in str(url):
            continue
        vid = str(url).split("watch?v=")[1].split("&")[0]
        title = r[1] or ""
        chan = r[3] or ""
        rows.append((vid, url, title, chan))
    # de-dupe by id, keep first
    seen, out = set(), []
    for v in rows:
        if v[0] in seen: continue
        seen.add(v[0]); out.append(v)
    return out

def clean(text):
    text = re.sub(r"\s+", " ", text).strip()
    return text

# ---- method 1: youtube-transcript-api --------------------------------------
def via_api(vid):
    from youtube_transcript_api import YouTubeTranscriptApi
    kwargs = {}
    if PROXY:
        from youtube_transcript_api.proxies import GenericProxyConfig
        kwargs["proxy_config"] = GenericProxyConfig(http_url=PROXY, https_url=PROXY)
    api = YouTubeTranscriptApi(**kwargs)
    # newer API: .fetch returns snippet objects with .text ; prefer English
    try:
        tr = api.fetch(vid, languages=["en", "en-US", "en-GB"])
    except TypeError:
        tr = api.fetch(vid)
    parts = [getattr(s, "text", s.get("text", "")) for s in tr]
    return clean(" ".join(parts))

# ---- method 2: yt-dlp auto subs --------------------------------------------
def via_ytdlp_subs(vid):
    import subprocess, tempfile, os
    with tempfile.TemporaryDirectory() as d:
        cmd = ["yt-dlp", "--skip-download", "--write-auto-sub", "--write-sub",
               "--sub-lang", "en.*", "--sub-format", "vtt",
               "-o", os.path.join(d, "%(id)s"),
               f"https://www.youtube.com/watch?v={vid}"]
        if PROXY: cmd += ["--proxy", PROXY]
        subprocess.run(cmd, capture_output=True, timeout=120)
        vtts = glob.glob(os.path.join(d, f"{vid}*.vtt"))
        if not vtts: return None
        return clean(vtt_to_text(open(vtts[0], encoding="utf-8").read()))

def vtt_to_text(vtt):
    lines = []
    for ln in vtt.splitlines():
        if "-->" in ln or ln.strip().isdigit() or ln.startswith("WEBVTT") or not ln.strip():
            continue
        ln = re.sub(r"<[^>]+>", "", ln)  # strip inline tags
        lines.append(ln)
    # collapse the duplicate lines auto-subs love to produce
    out = []
    for ln in lines:
        if not out or out[-1] != ln:
            out.append(ln)
    return " ".join(out)

# ---- method 3: whisper on audio --------------------------------------------
_whisper = None
def via_whisper(vid):
    import subprocess, tempfile, os
    global _whisper
    if _whisper is None:
        from faster_whisper import WhisperModel
        _whisper = WhisperModel(WHISPER_MODEL, device="auto", compute_type="auto")
    with tempfile.TemporaryDirectory() as d:
        audio = os.path.join(d, f"{vid}.m4a")
        cmd = ["yt-dlp", "-f", "bestaudio", "-x", "--audio-format", "m4a",
               "-o", os.path.join(d, "%(id)s.%(ext)s"),
               f"https://www.youtube.com/watch?v={vid}"]
        if PROXY: cmd += ["--proxy", PROXY]
        subprocess.run(cmd, capture_output=True, timeout=600)
        cand = glob.glob(os.path.join(d, f"{vid}.*"))
        if not cand: return None
        segs, _ = _whisper.transcribe(cand[0], language="en")
        return clean(" ".join(s.text for s in segs))

def attempt(fn, vid):
    last = None
    for i in range(RETRIES):
        try:
            t = fn(vid)
            if t: return t, None
            return None, "empty"
        except Exception as e:
            last = f"{type(e).__name__}: {str(e)[:120]}"
            time.sleep(2 ** i)
    return None, last

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--no-whisper", action="store_true")
    args = ap.parse_args()

    os.makedirs(OUT_DIR, exist_ok=True)
    ids = read_ids(XLSX_PATH)
    if args.limit: ids = ids[:args.limit]
    log(f"{len(ids)} videos to process")

    done = {os.path.splitext(os.path.basename(p))[0]
            for p in glob.glob(os.path.join(OUT_DIR, "*.txt"))}

    new_manifest = not os.path.exists(MANIFEST)
    mf = open(MANIFEST, "a", newline="", encoding="utf-8")
    w = csv.writer(mf)
    if new_manifest:
        w.writerow(["video_id","url","title","channel","status","source","chars"])

    counts = {"api":0,"ytdlp":0,"whisper":0,"none":0,"skip":0}
    for i,(vid,url,title,chan) in enumerate(ids,1):
        if vid in done:
            counts["skip"]+=1; continue

        text, src = None, None
        t,_ = attempt(via_api, vid)
        if t: text, src = t, "api"
        if not text:
            t,_ = attempt(via_ytdlp_subs, vid)
            if t: text, src = t, "ytdlp"
        if not text and not args.no_whisper:
            t,_ = attempt(via_whisper, vid)
            if t: text, src = t, "whisper"

        if text:
            with open(os.path.join(OUT_DIR, f"{vid}.txt"), "w", encoding="utf-8") as f:
                f.write(text)
            counts[src]+=1
            w.writerow([vid,url,title,chan,"ok",src,len(text)])
            log(f"[{i}/{len(ids)}] {vid} OK via {src} ({len(text)} chars)")
        else:
            counts["none"]+=1
            w.writerow([vid,url,title,chan,"no_transcript","",0])
            log(f"[{i}/{len(ids)}] {vid} NO TRANSCRIPT")
        mf.flush()
        time.sleep(random.uniform(*DELAY))

    mf.close()
    log("\nDONE:", counts)
    log(f"Transcripts in ./{OUT_DIR}/   |   status log in ./{MANIFEST}")

if __name__ == "__main__":
    main()
